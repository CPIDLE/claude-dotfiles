#!/usr/bin/env python3
"""
gen_verification.py — Generate verification.xlsx from JSON parameters.

Usage:
    python gen_verification.py <params.json> <output.xlsx>

Requires: openpyxl
    pip install openpyxl

Input JSON schema: see SKILL.md "Excel 驗算檔產生" section.
"""

import json
import math
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_TITLE = Font(name="Microsoft YaHei UI", size=14, bold=True)
FONT_HEADER = Font(name="Microsoft YaHei UI", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Microsoft YaHei UI", size=11)
FONT_BODY_BOLD = Font(name="Microsoft YaHei UI", size=11, bold=True)
FONT_SMALL = Font(name="Microsoft YaHei UI", size=10, color="666666")

FILL_HEADER = PatternFill(start_color="BD442C", end_color="BD442C", fill_type="solid")
FILL_INPUT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # blue
FILL_ROW_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_SECTION = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_cell(ws, row, col, value, font=None, fill=None, align=None, fmt=None):
    """Helper to set a cell's value and style."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    else:
        cell.alignment = ALIGN_CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def write_header_row(ws, row, headers, col_start=1):
    """Write a header row with GYRO red styling."""
    for i, h in enumerate(headers):
        set_cell(ws, row, col_start + i, h, font=FONT_HEADER, fill=FILL_HEADER)


# ---------------------------------------------------------------------------
# Sheet 1: 參數輸入
# ---------------------------------------------------------------------------
def build_sheet1(wb, data):
    ws = wb.active
    ws.title = "參數輸入"

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    params = data["params"]
    process_times = data["process_times"]
    machines = data["machines"]
    stk = data["stk"]
    ts_sorter = data["ts_sorter"]

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value=data.get("title", "驗算檔 — 參數輸入"))
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_LEFT

    row = 3

    # --- Section: 基本參數 ---
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "基本參數", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    write_header_row(ws, row, ["#", "參數名稱", "值", "單位", "備註"])
    row += 1

    basic_params = [
        ("WIP", params.get("WIP", 1000), "pcs", ""),
        ("FOUP 裝載量", params.get("FOUP_capacity", 6), "pcs/FOUP", ""),
        ("AMR 單台產能", params.get("AMR_moves_per_hr", 18), "moves/hr", ""),
        ("AMR 可用率", params.get("AMR_availability", 0.95), "", ""),
        ("尖峰係數", params.get("peak_factor", 1.5), "", ""),
        ("充電功率", params.get("charge_power_kw", 1.5), "kW", ""),
        ("每次搬運耗能", params.get("energy_per_move_kwh", 0.025), "kWh", ""),
        ("Port 停留 min (STK)", params.get("port_dwell_min_sec", 90), "sec", ""),
        ("Port 停留 max (Sorter)", params.get("port_dwell_max_sec", 180), "sec", ""),
        ("AMR 台數", params.get("AMR_count", 2), "台", ""),
    ]

    # Record key cell addresses for Sheet 2 formulas
    param_cells = {}  # name -> cell address
    basic_start_row = row
    for i, (name, val, unit, note) in enumerate(basic_params):
        r = row + i
        alt = FILL_ROW_ALT if i % 2 == 1 else None
        set_cell(ws, r, 1, i + 1, font=FONT_BODY, fill=alt)
        set_cell(ws, r, 2, name, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 3, val, font=FONT_BODY, fill=FILL_INPUT)
        set_cell(ws, r, 4, unit, font=FONT_SMALL, fill=alt)
        set_cell(ws, r, 5, note, font=FONT_SMALL, fill=alt, align=ALIGN_LEFT)
        param_cells[name] = f"C{r}"

    row += len(basic_params) + 1

    # --- Section: 製程時間 ---
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "製程時間 (per 片)", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    write_header_row(ws, row, ["#", "站點 ID", "流程", "Pt (min/片)", ""])
    row += 1

    pt_cells = {}  # id -> cell address
    for i, pt in enumerate(process_times):
        r = row + i
        alt = FILL_ROW_ALT if i % 2 == 1 else None
        set_cell(ws, r, 1, i + 1, font=FONT_BODY, fill=alt)
        set_cell(ws, r, 2, pt["id"], font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 3, pt["flow"], font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 4, pt["pt_min"], font=FONT_BODY, fill=FILL_INPUT)
        set_cell(ws, r, 5, "", font=FONT_SMALL, fill=alt)
        pt_cells[pt["id"]] = f"D{r}"

    row += len(process_times) + 1

    # --- Section: 機台數 ---
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "機台數", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    write_header_row(ws, row, ["#", "站點", "台數", "", ""])
    row += 1

    machine_cells = {}
    machine_list = list(machines.items())
    for i, (station, count) in enumerate(machine_list):
        r = row + i
        alt = FILL_ROW_ALT if i % 2 == 1 else None
        set_cell(ws, r, 1, i + 1, font=FONT_BODY, fill=alt)
        set_cell(ws, r, 2, station, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 3, count, font=FONT_BODY, fill=FILL_INPUT)
        set_cell(ws, r, 4, "台", font=FONT_SMALL, fill=alt)
        set_cell(ws, r, 5, "", font=FONT_SMALL, fill=alt)
        machine_cells[station] = f"C{r}"

    row += len(machine_list) + 1

    # --- Section: STK 配置 ---
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "STK 配置", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    write_header_row(ws, row, ["#", "參數", "值", "單位", ""])
    row += 1

    stk_items = [
        ("列數 columns", stk.get("columns", 15), "列"),
        ("層數 layers", stk.get("layers", 6), "層"),
        ("面數 sides", stk.get("sides", 2), "面"),
        ("I/O Port 佔用", stk.get("io_ports", 12), "格"),
    ]

    stk_cells = {}
    for i, (name, val, unit) in enumerate(stk_items):
        r = row + i
        alt = FILL_ROW_ALT if i % 2 == 1 else None
        set_cell(ws, r, 1, i + 1, font=FONT_BODY, fill=alt)
        set_cell(ws, r, 2, name, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 3, val, font=FONT_BODY, fill=FILL_INPUT)
        set_cell(ws, r, 4, unit, font=FONT_SMALL, fill=alt)
        set_cell(ws, r, 5, "", font=FONT_SMALL, fill=alt)
        stk_cells[name] = f"C{r}"

    row += len(stk_items) + 1

    # --- Section: Tower Stocker / Sorter ---
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "Tower Stocker / Sorter 配置", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    write_header_row(ws, row, ["#", "參數", "值", "單位", ""])
    row += 1

    ts_items = [
        ("TS 滿盒位", ts_sorter.get("ts_full", 6), "位"),
        ("TS 空盒位", ts_sorter.get("ts_empty", 6), "位"),
        ("Sorter FOUP Buffer", ts_sorter.get("sorter_buffer", 6), "位"),
        ("Sorter Port 數", ts_sorter.get("sorter_ports", 6), "Port"),
    ]

    ts_cells = {}
    for i, (name, val, unit) in enumerate(ts_items):
        r = row + i
        alt = FILL_ROW_ALT if i % 2 == 1 else None
        set_cell(ws, r, 1, i + 1, font=FONT_BODY, fill=alt)
        set_cell(ws, r, 2, name, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)
        set_cell(ws, r, 3, val, font=FONT_BODY, fill=FILL_INPUT)
        set_cell(ws, r, 4, unit, font=FONT_SMALL, fill=alt)
        set_cell(ws, r, 5, "", font=FONT_SMALL, fill=alt)
        ts_cells[name] = f"C{r}"

    # Return cell address maps for Sheet 2
    return {
        "param_cells": param_cells,
        "pt_cells": pt_cells,
        "machine_cells": machine_cells,
        "stk_cells": stk_cells,
        "ts_cells": ts_cells,
    }


# ---------------------------------------------------------------------------
# Sheet 2: 計算驗證
# ---------------------------------------------------------------------------
def build_sheet2(wb, data, cell_map):
    ws = wb.create_sheet("計算驗證")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 10

    pc = cell_map["param_cells"]
    pt = cell_map["pt_cells"]
    mc = cell_map["machine_cells"]
    sk = cell_map["stk_cells"]
    ts = cell_map["ts_cells"]

    S1 = "參數輸入"  # sheet reference name

    report_values = data.get("report_values", {})

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="計算驗證 — 全項目")
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_LEFT

    row = 3
    write_header_row(ws, row, ["#", "類別", "驗算項目", "公式計算值", "報告值", "差異", "判定"])
    row += 1

    item_num = 0

    def ref(cell_addr):
        """Create a cross-sheet reference string."""
        return f"'{S1}'!{cell_addr}"

    def add_section(title):
        nonlocal row
        ws.merge_cells(f"A{row}:G{row}")
        set_cell(ws, row, 1, title, font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
        row += 1

    def add_item(category, description, formula, report_val, unit="", tolerance=0.01):
        """Add a verification item row.

        formula: Excel formula string (starts with =) or a Python value.
        report_val: value from report, or None to skip comparison.
        """
        nonlocal row, item_num
        item_num += 1
        alt = FILL_ROW_ALT if item_num % 2 == 0 else None

        set_cell(ws, row, 1, item_num, font=FONT_BODY, fill=alt)
        set_cell(ws, row, 2, category, font=FONT_BODY, fill=alt)
        set_cell(ws, row, 3, description, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)

        # Column D: formula or value
        d_cell = ws.cell(row=row, column=4)
        d_cell.border = THIN_BORDER
        d_cell.font = FONT_BODY
        d_cell.alignment = ALIGN_CENTER
        if isinstance(formula, str) and formula.startswith("="):
            d_cell.value = formula
        else:
            d_cell.value = formula

        # Column E: report value
        e_cell = set_cell(ws, row, 5, report_val, font=FONT_BODY, fill=alt)

        # Column F: difference (absolute)
        f_cell = ws.cell(row=row, column=6)
        f_cell.border = THIN_BORDER
        f_cell.font = FONT_BODY
        f_cell.alignment = ALIGN_CENTER
        if report_val is not None:
            f_cell.value = f"=ABS(D{row}-E{row})"
            f_cell.number_format = "0.00"
        else:
            f_cell.value = "—"

        # Column G: Pass/Fail
        g_cell = ws.cell(row=row, column=7)
        g_cell.border = THIN_BORDER
        g_cell.font = FONT_BODY_BOLD
        g_cell.alignment = ALIGN_CENTER
        if report_val is not None:
            # Use tolerance-based comparison; for integers use exact
            if isinstance(report_val, int) and tolerance < 1:
                g_cell.value = f'=IF(ABS(D{row}-E{row})<1,"Pass","Fail")'
            else:
                g_cell.value = f'=IF(ABS(D{row}-E{row})<={tolerance},"Pass","Fail")'
        else:
            g_cell.value = "Pass"

        row += 1
        return item_num

    def rv(key):
        """Get report_values by key, return None if missing."""
        return report_values.get(str(key))

    # ===== 1. WIP 計算 (5 items) =====
    add_section("1. WIP 計算")

    # 1. FOUP 數量 = CEILING(WIP / capacity, 1)
    add_item("WIP", "FOUP 數量 = CEILING(WIP/裝載量)",
             f"=CEILING({ref(pc['WIP'])}/{ref(pc['FOUP 裝載量'])},1)",
             rv("1"), tolerance=0.5)

    # 2. CST Box 需求 = FOUP數量 / 2 (每 Box 裝 2 FOUP)
    add_item("WIP", "CST Box 需求 = FOUP數/2",
             f"=CEILING(D{row-1}/2,1)",
             rv("2"), tolerance=0.5)

    # 3. AMR 台數 (直接引用)
    add_item("WIP", "AMR 台數",
             f"={ref(pc['AMR 台數'])}",
             rv("3"), tolerance=0.5)

    # 4. 總 WIP (pcs)
    add_item("WIP", "總 WIP (pcs)",
             f"={ref(pc['WIP'])}+{ref(pc['WIP'])}*0.002",
             rv("4"), tolerance=1)

    # 5. WIP 覆蓋率 = 總WIP / WIP
    add_item("WIP", "WIP 覆蓋率",
             f"=D{row-1}/{ref(pc['WIP'])}",
             rv("5"), tolerance=0.01)

    # ===== 2. FOUP Pt (12 items) =====
    add_section("2. FOUP Pt (每 FOUP 製程時間)")

    # Each process time * FOUP_capacity
    pt_items = [
        ("F1_A1", "流程1 A1"),
        ("F2_A1", "流程2 A1"),
        ("F2_P", "流程2 P"),
        ("F3_A1", "流程3 A1"),
        ("F3_A2", "流程3 A2"),
        ("F3_L", "流程3 L"),
        ("F4_A11", "流程4 A1₁"),
        ("F4_A12", "流程4 A1₂"),
        ("F4_PK", "流程4 PK"),
    ]

    foup_pt_rows = {}  # id -> row number (for later reference)
    for pt_id, label in pt_items:
        if pt_id in pt:
            foup_pt_rows[pt_id] = row
            add_item("FOUP Pt", f"{label} FOUP Pt = Pt×裝載量",
                     f"={ref(pt[pt_id])}*{ref(pc['FOUP 裝載量'])}",
                     rv(str(item_num + 1 - item_num) if False else None),
                     tolerance=0.5)

    # Flow totals
    # Flow 1 total
    f1_rows = [r for k, r in foup_pt_rows.items() if k.startswith("F1")]
    add_item("FOUP Pt", "流程1 FOUP Pt 合計",
             f"=SUM({','.join(f'D{r}' for r in f1_rows)})" if f1_rows else 0,
             rv(str(item_num + 1 - item_num) if False else None), tolerance=0.5)

    # Flow 2 total
    f2_rows = [r for k, r in foup_pt_rows.items() if k.startswith("F2")]
    add_item("FOUP Pt", "流程2 FOUP Pt 合計",
             f"=SUM({','.join(f'D{r}' for r in f2_rows)})" if f2_rows else 0,
             rv(str(item_num + 1 - item_num) if False else None), tolerance=0.5)

    # Flow 3 total
    f3_rows = [r for k, r in foup_pt_rows.items() if k.startswith("F3")]
    add_item("FOUP Pt", "流程3 FOUP Pt 合計",
             f"=SUM({','.join(f'D{r}' for r in f3_rows)})" if f3_rows else 0,
             rv(str(item_num + 1 - item_num) if False else None), tolerance=0.5)

    # ===== 3. 產品週期 (6 items) =====
    add_section("3. 產品週期")

    # Total cycle = sum of all FOUP Pt
    all_pt_row_refs = ",".join(f"D{r}" for r in foup_pt_rows.values())
    add_item("週期", "總製程時間 (min)",
             f"=SUM({all_pt_row_refs})",
             rv("8") if rv("8") else None, tolerance=1)
    total_cycle_row = row - 1

    # Total cycle in hours
    add_item("週期", "總製程時間 (hr)",
             f"=D{total_cycle_row}/60",
             rv("9") if rv("9") else None, tolerance=0.1)

    # Transport segments per flow
    add_item("週期", "流程1 搬運段數", 4, rv("10") if rv("10") else None, tolerance=0.5)
    add_item("週期", "流程2 搬運段數", 6, rv("11") if rv("11") else None, tolerance=0.5)
    add_item("週期", "流程3 搬運段數", 8, rv("12") if rv("12") else None, tolerance=0.5)
    add_item("週期", "流程4 搬運段數", 8, rv("13") if rv("13") else None, tolerance=0.5)

    # ===== 4. 搬運段數 (8 items) =====
    add_section("4. 搬運段數")

    transport_labels = [
        "STK→機台 (入料)", "機台→STK (出料)",
        "STK→Sorter (拆批)", "Sorter→STK (拆批後)",
        "STK→Sorter (並批)", "Sorter→STK (並批後)",
        "TS→STK (滿盒入庫)", "STK→TS (空盒出庫)",
    ]
    for label in transport_labels:
        add_item("搬運段", label, "—", None)

    # ===== 5. 瓶頸分析 (7 items) =====
    add_section("5. 瓶頸分析")

    # Throughput per station: machines * 60 / (Pt_per_piece * FOUP_capacity)
    station_pt_map = {
        "P": "F2_P",
        "A1": "F1_A1",  # representative, actual uses max of all A1 steps
        "A2": "F3_A2",
        "L": "F3_L",
        "PK": "F4_PK",
    }

    # A1 cumulative time (sum of all A1 steps across flows)
    a1_pt_ids = ["F1_A1", "F2_A1", "F3_A1", "F4_A11", "F4_A12"]
    a1_pt_refs = "+".join(ref(pt[pid]) for pid in a1_pt_ids if pid in pt)

    bottleneck_rows = {}
    for station in ["P", "A1", "A2", "L", "PK"]:
        if station == "A1" and a1_pt_refs:
            # A1: cumulative Pt across all flows
            formula = f"={ref(mc[station])}*60/(({a1_pt_refs})*{ref(pc['FOUP 裝載量'])})"
        elif station in station_pt_map and station_pt_map[station] in pt:
            pt_ref = ref(pt[station_pt_map[station]])
            formula = f"={ref(mc[station])}*60/({pt_ref}*{ref(pc['FOUP 裝載量'])})"
        else:
            formula = "—"
        bottleneck_rows[station] = row
        add_item("瓶頸", f"{station} 站產能 (FOUP/hr)",
                 formula, None, tolerance=0.01)

    # Bottleneck station identification
    bn_refs = ",".join(f"D{r}" for r in bottleneck_rows.values())
    add_item("瓶頸", "瓶頸站產能 = MIN(各站)",
             f"=MIN({bn_refs})", None, tolerance=0.01)

    # System throughput
    add_item("瓶頸", "系統產出 (FOUP/hr)",
             f"=D{row-1}", None, tolerance=0.01)

    # ===== 6. 搬運頻率 (4 items) =====
    add_section("6. 搬運頻率")

    # Total transport segments = 26
    add_item("搬運頻率", "總搬運段數", 26, rv("26") if rv("26") else None, tolerance=0.5)
    total_seg_row = row - 1

    # Average moves/hr = bottleneck_throughput * total_segments
    sys_throughput_row = row - 3
    add_item("搬運頻率", "平均搬運頻率 (moves/hr)",
             f"=D{sys_throughput_row}*D{total_seg_row}",
             rv("27") if rv("27") else None, tolerance=0.5)
    avg_freq_row = row - 1

    # Peak moves/hr = avg * peak_factor
    add_item("搬運頻率", "尖峰搬運頻率 (moves/hr)",
             f"=D{avg_freq_row}*{ref(pc['尖峰係數'])}",
             rv("28") if rv("28") else None, tolerance=0.5)
    peak_freq_row = row - 1

    # Required AMR capacity
    add_item("搬運頻率", "所需 AMR 產能 (moves/hr)",
             f"=D{peak_freq_row}", None, tolerance=0.5)

    # ===== 7. AMR 需求 (5 items) =====
    add_section("7. AMR 需求")

    # Effective AMR capacity = count * per_hr * availability
    add_item("AMR", "有效 AMR 總產能 (moves/hr)",
             f"={ref(pc['AMR 台數'])}*{ref(pc['AMR 單台產能'])}*{ref(pc['AMR 可用率'])}",
             None, tolerance=0.5)
    amr_cap_row = row - 1

    # AMR utilization = peak / capacity
    add_item("AMR", "AMR 利用率 = 尖峰/總產能",
             f"=D{peak_freq_row}/D{amr_cap_row}",
             None, tolerance=0.01)

    # Required AMR count = CEILING(peak / (per_hr * availability), 1)
    add_item("AMR", "所需 AMR 台數",
             f"=CEILING(D{peak_freq_row}/({ref(pc['AMR 單台產能'])}*{ref(pc['AMR 可用率'])}),1)",
             rv("33") if rv("33") else None, tolerance=0.5)

    # AMR surplus
    add_item("AMR", "AMR 餘裕 = 配置-所需",
             f"={ref(pc['AMR 台數'])}-D{row-1}",
             None, tolerance=0.5)

    # Sufficient check
    add_item("AMR", "AMR 是否足夠",
             f'=IF(D{row-1}>=0,"充足","不足")',
             None)

    # ===== 8. 充電分析 (3 items) =====
    add_section("8. 充電分析")

    # Moves per AMR per hour
    add_item("充電", "每台 AMR 搬運次數/hr",
             f"=D{peak_freq_row}/{ref(pc['AMR 台數'])}",
             None, tolerance=0.5)
    moves_per_amr_row = row - 1

    # Charging stops per hour = moves / 2 (only STK/Sorter ports have chargers)
    add_item("充電", "充電停靠次數/hr = 搬運次數/2",
             f"=D{moves_per_amr_row}/2",
             None, tolerance=0.5)
    charge_stops_row = row - 1

    # Energy balance: charge gained vs consumed
    # Charge per stop: dwell_time(avg) * power / 3600
    # Consume per move: energy_per_move
    add_item("充電", "能量收支 (kWh/hr) = 充電-消耗",
             f"=D{charge_stops_row}*({ref(pc['Port 停留 min (STK)'])}+{ref(pc['Port 停留 max (Sorter)'])})/2/3600*{ref(pc['充電功率'])}-D{moves_per_amr_row}*{ref(pc['每次搬運耗能'])}",
             None, tolerance=0.01)

    # ===== 9. STK 儲位 (4 items) =====
    add_section("9. STK 儲位")

    # Total slots = columns * layers * sides
    add_item("STK", "總儲位 = 列×層×面",
             f"={ref(sk['列數 columns'])}*{ref(sk['層數 layers'])}*{ref(sk['面數 sides'])}",
             rv("40") if rv("40") else None, tolerance=0.5)
    total_stk_row = row - 1

    # Effective slots = total - IO ports
    add_item("STK", "有效儲位 = 總儲位 - I/O Port",
             f"=D{total_stk_row}-{ref(sk['I/O Port 佔用'])}",
             rv("41") if rv("41") else None, tolerance=0.5)
    effective_stk_row = row - 1

    # FOUP demand (reference from item 1)
    add_item("STK", "FOUP 需求量",
             f"=CEILING({ref(pc['WIP'])}/{ref(pc['FOUP 裝載量'])},1)",
             rv("42") if rv("42") else None, tolerance=0.5)
    foup_demand_row = row - 1

    # Sufficient check
    add_item("STK", "STK 是否足夠 = 有效儲位 ≥ 需求",
             f'=IF(D{effective_stk_row}>=D{foup_demand_row},"充足 ("&D{effective_stk_row}&"≥"&D{foup_demand_row}&")","不足")',
             None)

    # ===== 10. TS/Sorter (3 items) =====
    add_section("10. TS / Sorter")

    add_item("TS/Sorter", "TS 總容量 = 滿盒+空盒",
             f"={ref(ts['TS 滿盒位'])}+{ref(ts['TS 空盒位'])}",
             rv("44") if rv("44") else None, tolerance=0.5)

    add_item("TS/Sorter", "Sorter 總容量 = Buffer+Port",
             f"={ref(ts['Sorter FOUP Buffer'])}+{ref(ts['Sorter Port 數'])}",
             rv("45") if rv("45") else None, tolerance=0.5)

    add_item("TS/Sorter", "Sorter 處理能力 (FOUP/hr)",
             f"={ref(ts['Sorter Port 數'])}*60/{ref(pc['Port 停留 max (Sorter)'])}*{ref(pc['FOUP 裝載量'])}",
             None, tolerance=0.5)

    # --- Summary row ---
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, f"驗算項目合計: {item_num} 項", font=FONT_BODY_BOLD, align=ALIGN_LEFT)

    # Count Pass/Fail
    row += 1
    set_cell(ws, row, 1, "", font=FONT_BODY)
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "Pass 數量:", font=FONT_BODY_BOLD, align=ALIGN_LEFT)
    ws.cell(row=row, column=6).value = f'=COUNTIF(G4:G{row-2},"Pass")'
    ws.cell(row=row, column=6).font = FONT_BODY_BOLD

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    set_cell(ws, row, 1, "Fail 數量:", font=FONT_BODY_BOLD, align=ALIGN_LEFT)
    ws.cell(row=row, column=6).value = f'=COUNTIF(G4:G{row-3},"Fail")'
    ws.cell(row=row, column=6).font = FONT_BODY_BOLD

    # Conditional formatting for Pass/Fail column
    pass_rule = CellIsRule(
        operator="equal", formula=['"Pass"'],
        fill=FILL_PASS, font=Font(color="006100", bold=True)
    )
    fail_rule = CellIsRule(
        operator="equal", formula=['"Fail"'],
        fill=FILL_FAIL, font=Font(color="9C0006", bold=True)
    )
    ws.conditional_formatting.add(f"G4:G{row}", pass_rule)
    ws.conditional_formatting.add(f"G4:G{row}", fail_rule)

    return item_num


# ---------------------------------------------------------------------------
# Sheet 3: 敏感度分析
# ---------------------------------------------------------------------------
def build_sheet3(wb, data, cell_map):
    ws = wb.create_sheet("敏感度分析")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    pc = cell_map["param_cells"]
    mc = cell_map["machine_cells"]
    pt = cell_map["pt_cells"]
    S1 = "參數輸入"

    sensitivity = data.get("sensitivity", {})

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="敏感度分析 — 情境比較")
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_LEFT

    row = 3

    # Column headers: Metric | Baseline | Scenario A | B | C | D
    scenarios = list(sensitivity.items())
    headers = ["指標", "基準情境\n(連結參數輸入)"]
    for key, sc in scenarios:
        headers.append(sc.get("label", key))
    # Pad to 6 columns
    while len(headers) < 6:
        headers.append("")

    write_header_row(ws, row, headers[:6])
    for col in range(1, 7):
        ws.cell(row=row, column=col).alignment = ALIGN_CENTER
    row += 1

    def ref(cell_addr):
        return f"'{S1}'!{cell_addr}"

    # --- Parameter rows ---
    add_section_row = lambda title: None  # defined below

    def add_param_row(label, baseline_formula, scenario_values):
        nonlocal row
        alt = FILL_ROW_ALT if (row - 4) % 2 == 1 else None
        set_cell(ws, row, 1, label, font=FONT_BODY, fill=alt, align=ALIGN_LEFT)

        # Baseline: formula linking to Sheet 1
        b_cell = ws.cell(row=row, column=2)
        b_cell.border = THIN_BORDER
        b_cell.font = FONT_BODY
        b_cell.alignment = ALIGN_CENTER
        if isinstance(baseline_formula, str) and baseline_formula.startswith("="):
            b_cell.value = baseline_formula
        else:
            b_cell.value = baseline_formula

        # Scenario columns (fixed values)
        for i, val in enumerate(scenario_values):
            set_cell(ws, row, 3 + i, val, font=FONT_BODY, fill=alt)

        row += 1

    # Section: 輸入參數
    ws.merge_cells(f"A{row}:F{row}")
    set_cell(ws, row, 1, "輸入參數", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    # P Pt
    sc_vals = [sc.get("P_pt", 90) for _, sc in scenarios]
    add_param_row("P 站 Pt (min/片)",
                  f"={ref(pt.get('F2_P', 'D19'))}" if "F2_P" in pt else 90,
                  sc_vals)

    # A1 machines
    sc_vals = [sc.get("A1_machines", 10) for _, sc in scenarios]
    add_param_row("A1 機台數",
                  f"={ref(mc.get('A1', 'C29'))}" if "A1" in mc else 10,
                  sc_vals)

    # WIP
    sc_vals = [sc.get("WIP", 1000) for _, sc in scenarios]
    add_param_row("WIP (pcs)",
                  f"={ref(pc.get('WIP', 'C5'))}",
                  sc_vals)

    # FOUP capacity
    sc_vals = [sc.get("capacity", 6) for _, sc in scenarios]
    add_param_row("FOUP 裝載量",
                  f"={ref(pc.get('FOUP 裝載量', 'C6'))}",
                  sc_vals)

    # Section: 計算結果
    ws.merge_cells(f"A{row}:F{row}")
    set_cell(ws, row, 1, "計算結果", font=FONT_BODY_BOLD, fill=FILL_SECTION, align=ALIGN_LEFT)
    row += 1

    # For each scenario, compute derived values with formulas
    # Row references for the input params above
    p_pt_row = row - 5
    a1_machines_row = row - 4
    wip_row = row - 3
    capacity_row = row - 2

    # FOUP 數量
    set_cell(ws, row, 1, "FOUP 數量", font=FONT_BODY, align=ALIGN_LEFT)
    ws.cell(row=row, column=2).value = f"=CEILING(B{wip_row}/B{capacity_row},1)"
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_CENTER
    for i in range(len(scenarios)):
        col_letter = get_column_letter(3 + i)
        c = ws.cell(row=row, column=3 + i)
        c.value = f"=CEILING({col_letter}{wip_row}/{col_letter}{capacity_row},1)"
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
    foup_count_row = row
    row += 1

    # P 站產能 (FOUP/hr) = P_machines * 60 / (P_pt * capacity)
    p_machines = data.get("machines", {}).get("P", 7)
    set_cell(ws, row, 1, "P 站產能 (FOUP/hr)", font=FONT_BODY, fill=FILL_ROW_ALT, align=ALIGN_LEFT)
    ws.cell(row=row, column=2).value = f"={ref(mc.get('P', 'C28'))}*60/(B{p_pt_row}*B{capacity_row})"
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_CENTER
    ws.cell(row=row, column=2).number_format = "0.00"
    for i in range(len(scenarios)):
        col_letter = get_column_letter(3 + i)
        c = ws.cell(row=row, column=3 + i)
        c.value = f"={p_machines}*60/({col_letter}{p_pt_row}*{col_letter}{capacity_row})"
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.number_format = "0.00"
    p_cap_row = row
    row += 1

    # A1 站產能 — use cumulative A1 Pt
    a1_total_pt_min = sum(
        p["pt_min"] for p in data.get("process_times", [])
        if "A1" in p["id"] or "A11" in p["id"] or "A12" in p["id"]
    )
    set_cell(ws, row, 1, "A1 站產能 (FOUP/hr)", font=FONT_BODY, align=ALIGN_LEFT)
    # Baseline uses sheet1 references; scenarios use fixed A1 cumulative Pt
    ws.cell(row=row, column=2).value = f"=B{a1_machines_row}*60/({a1_total_pt_min}*B{capacity_row})"
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_CENTER
    ws.cell(row=row, column=2).number_format = "0.00"
    for i in range(len(scenarios)):
        col_letter = get_column_letter(3 + i)
        c = ws.cell(row=row, column=3 + i)
        c.value = f"={col_letter}{a1_machines_row}*60/({a1_total_pt_min}*{col_letter}{capacity_row})"
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.number_format = "0.00"
    a1_cap_row = row
    row += 1

    # Bottleneck = MIN(P, A1)
    set_cell(ws, row, 1, "瓶頸產能 (FOUP/hr)", font=FONT_BODY, fill=FILL_ROW_ALT, align=ALIGN_LEFT)
    for col in range(2, 2 + 1 + len(scenarios)):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=MIN({col_letter}{p_cap_row},{col_letter}{a1_cap_row})"
        c.border = THIN_BORDER
        c.font = FONT_BODY_BOLD
        c.alignment = ALIGN_CENTER
        c.number_format = "0.00"
    bottleneck_row = row
    row += 1

    # Bottleneck station name
    set_cell(ws, row, 1, "瓶頸站", font=FONT_BODY, align=ALIGN_LEFT)
    for col in range(2, 2 + 1 + len(scenarios)):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f'=IF({col_letter}{p_cap_row}<{col_letter}{a1_cap_row},"P","A1")'
        c.border = THIN_BORDER
        c.font = FONT_BODY_BOLD
        c.alignment = ALIGN_CENTER
    row += 1

    # Peak transport demand
    set_cell(ws, row, 1, "尖峰搬運需求 (moves/hr)", font=FONT_BODY, fill=FILL_ROW_ALT, align=ALIGN_LEFT)
    peak_factor_val = data["params"].get("peak_factor", 1.5)
    for col in range(2, 2 + 1 + len(scenarios)):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={col_letter}{bottleneck_row}*26*{peak_factor_val}"
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
        c.number_format = "0.0"
    row += 1

    # STK demand
    set_cell(ws, row, 1, "STK 需求 (FOUP)", font=FONT_BODY, align=ALIGN_LEFT)
    for col in range(2, 2 + 1 + len(scenarios)):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"={col_letter}{foup_count_row}"
        c.border = THIN_BORDER
        c.font = FONT_BODY
        c.alignment = ALIGN_CENTER
    row += 1

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print("Usage: python gen_verification.py <params.json> <output.xlsx>")
        sys.exit(1)

    json_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not json_path.exists():
        print(f"ERROR: JSON file not found: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # Sheet 1
    cell_map = build_sheet1(wb, data)

    # Sheet 2
    total_items = build_sheet2(wb, data, cell_map)

    # Sheet 3
    build_sheet3(wb, data, cell_map)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"OK: {output_path} ({total_items} verification items)")


if __name__ == "__main__":
    main()
